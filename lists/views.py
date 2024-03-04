import io
import json
import locale
import calendar
import datetime
import openpyxl

from openpyxl.styles import PatternFill, Color

from users.models import User
from .models import Paesu_Record
from django.http import FileResponse,HttpResponse
from django.core.files.storage import FileSystemStorage
from django.shortcuts import render, redirect

from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView


def ListView(request):

    if request.user.is_authenticated:

        try:
            year = int(request.GET['year'])
            month = int(request.GET['month'])
        except:
            today = datetime.datetime.today()
            year = today.year
            month = today.month

        if request.user.level == '2':
            start_date = datetime.datetime(year, month, 1)
            days_in_month = 31 if month == 12 else (datetime.datetime(year, month+1, 1) - datetime.datetime(year, month, 1)).days
            date_list = [start_date + datetime.timedelta(days=x) for x in range(days_in_month)]
            date_str_list = [d.strftime("%Y-%m-%d") for d in date_list]

            result = []
            for i in date_str_list:
                try:
                    psrecord = Paesu_Record.objects.filter(user_id_p = request.user).get(date = i)
                    result.append([i,psrecord.diswaste_today,psrecord.diswaste_used,'입력완료'])
                except:
                    result.append([i,'','','입력필요'])

            context = {"list" : result,}
                    #    "start_date" : start_date,
                    #    "end_date" : end_date}

            return render(request, 'lists/list.html', context)
        
        else:
            try:
                start_date = datetime.datetime.strptime(request.GET['start'], "%Y-%m-%d")
                end_date = datetime.datetime.strptime(request.GET['end'], "%Y-%m-%d")
                date_list = []
                delta = datetime.timedelta(days=1)
                while start_date <= end_date:
                    date_list.append(start_date.date())
                    start_date += delta
                date_str_list = [d.strftime("%Y-%m-%d") for d in date_list]

                selected= request.GET['filtered-select']

                result = []
                for i in date_str_list:
                    try:
                        psrecord = Paesu_Record.objects.filter(user_id_p = User.objects.get(business_name = selected)).get(date = i)
                        result.append([i,psrecord.diswaste_today,psrecord.diswaste_used,'입력완료'])
                    except:
                        result.append([i,'','','입력필요'])

                if request.user.level == '1':
                    all_list = User.objects.filter(level = 2, region = request.user.region)
                    my_list = [usr.business_name for usr in all_list]
                    context = {
                        "corp" : json.dumps(my_list),
                        "list" : result,
                        "start_date" : request.GET['start'],
                        "end_date" : request.GET['end'],
                        "selected_corp" : selected,
                    }
                
                    return render(request, 'lists/list.html', context)
                
                elif request.user.level == '0':

                    all_list = User.objects.filter(level = 2)
                    my_list = [usr.business_name for usr in all_list]

                    context = {
                        "corp" : json.dumps(my_list),
                        "list" : result,
                        "start_date" : request.GET['start'],
                        "end_date" : request.GET['end'],
                        "selected_corp" : selected,
                    }

                    return render(request, 'lists/list.html', context)

            except:

                if request.user.level == '1':
                    all_list = User.objects.filter(level = 2, region = request.user.region)
                    my_list = [usr.business_name for usr in all_list]
                    context = {
                        "corp" : json.dumps(my_list)
                    }
                
                    return render(request, 'lists/list.html', context)
                
                elif request.user.level == '0':

                    all_list = User.objects.filter(level = 2)
                    my_list = [usr.business_name for usr in all_list]

                    context = {
                        "corp" : json.dumps(my_list)
                    }

                    return render(request, 'lists/list.html', context)

    else:
        return redirect('/')
    

def InsertData(request, date):

    # page url 형식 맞추기"
    try:
        datetime.datetime.strptime(date, '%Y-%m-%d')
    except:
        return redirect('/insert/'+str(datetime.date.today()))
    

    if request.user.is_authenticated:

        if request.method == 'POST':

            record = request.POST

            rest = 'off'
            if record.get('rest'):
                rest = record['rest']

            updated_values = {
                'date_ck' : rest,
                'date_weather' : record.get('weather'),
                'date_temperature' : record.get('temperature'),

                'waterworks_prevd' : record.get('waterworks_prevd'),
                'waterworks_used' : record.get('waterworks_used'),
                'waterworks_today' : record.get('waterworks_today'),

                'underwater_prevd' : record.get('underwater_prevd'),
                'underwater_used' : record.get('underwater_used'),
                'underwater_today' : record.get('underwater_today'),

                'diswaste_prevd' : record.get('diswaste_prevd'),
                'diswaste_used' : record.get('diswaste_used'),
                'diswaste_today' : record.get('diswaste_today'),

                'poweruse_prevd' : record.get('poweruse_prevd'),
                'poweruse_used' : record.get('poweruse_used'),
                'poweruse_today' : record.get('poweruse_today'),
                'poweruse_start' : record.get('time1'),
                'poweruse_end' : record.get('time2'),
                'poweruse_etc' : record.get('poweruse_etc'),
                
                'genwaster' : record.get('genwaster'),
                'reuse' : record.get('reuse'),

                'at_washnum' : record.get('at_washnum'),
                'at_detergent' : record.get('at_detergent'),
                'at_detergent_use' : record.get('at_detergent_use'),
                'at_wax' : record.get('at_wax'),
                'at_wax_use' : record.get('at_wax_use'),
                'at_pom' : record.get('at_pom'),
                'at_pom_use' : record.get('at_pom_use'),
                'at_sub1' : record.get('at_sub1'),
                'at_sub1_memo' : record.get('at_sub1_memo'),
                'at_sub1_use' : record.get('at_sub1_use'),
                'at_sub2' : record.get('at_sub2'),
                'at_sub2_memo' : record.get('at_sub2_memo'),
                'at_sub2_use' : record.get('at_sub2_use'),
                'at_sub3' : record.get('at_sub3'),
                'at_sub3_memo' : record.get('at_sub3_memo'),
                'at_sub3_use' : record.get('at_sub3_use'),

                'op_start' : record.get('op_start'),
                'op_end' : record.get('op_end'),

                'emission_start' : record.get('emission_start'),
                'emission_end' : record.get('emission_end'),

                'prev_start' : record.get('prev_start'),
                'prev_end' : record.get('prev_end'),

                'med1_name' : record.get('med1_name'),
                'med1_used' : record.get('med1_used'),
                'med1_buy' : record.get('med1_buy'),
                'med1_balance' : record.get('med1_balance'),
                'med1_etc' : record.get('med1_etc'),

                'med2_name' : record.get('med2_name'),
                'med2_used' : record.get('med2_used'),
                'med2_buy' : record.get('med2_buy'),
                'med2_balance' : record.get('med2_balance'),
                'med2_etc' : record.get('med2_etc'),

                'med3_name' : record.get('med3_name'),
                'med3_used' : record.get('med3_used'),
                'med3_buy' : record.get('med3_buy'),
                'med3_balance' : record.get('med3_balance'),
                'med3_etc' : record.get('med3_etc'),

                'med4_name' : record.get('med4_name'),
                'med4_used' : record.get('med4_used'),
                'med4_buy' : record.get('med4_buy'),
                'med4_balance' : record.get('med4_balance'),
                'med4_etc' : record.get('med4_etc'),

                'sluge_gene' : record.get('sluge_gene'),
                'sluge_used' : record.get('sluge_used'),
                'sluge_keep' : record.get('sluge_keep'),
                'sluge_func' : record.get('sluge_func'),
                'sluge_place' : record.get('sluge_place'),
                'sluge_selfplace' : record.get('sluge_selfplace'),
                'sluge_corp' : record.get('sluge_corp'),

                'remarks' : record.get('remarks'),
                'advise' : record.get('advise')
            }

            Paesu_Record.objects.update_or_create(
                user_id_p = request.user,
                date = record['date'],

                defaults = updated_values
                )

            return redirect('/list')

        else:

            try:
                yesterday = (datetime.datetime.strptime(date, '%Y-%m-%d') - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
                
                psrecord = Paesu_Record.objects.filter(user_id_p = request.user).filter(date = date)


                waterworks_prevd = Paesu_Record.objects.filter(user_id_p = request.user).get(date = yesterday).waterworks_today if Paesu_Record.objects.filter(user_id_p = request.user).filter(date = yesterday).exists() else 0
                waterworks_prevd = 0 if waterworks_prevd is None else waterworks_prevd

                underwater_prevd = Paesu_Record.objects.filter(user_id_p = request.user).get(date = yesterday).underwater_today if Paesu_Record.objects.filter(user_id_p = request.user).filter(date = yesterday).exists() else 0
                underwater_prevd = 0 if underwater_prevd is None else underwater_prevd

                diswaste_prevd = Paesu_Record.objects.filter(user_id_p = request.user).get(date = yesterday).diswaste_today if Paesu_Record.objects.filter(user_id_p = request.user).filter(date = yesterday).exists() else 0
                diswaste_prevd = 0 if diswaste_prevd is None else diswaste_prevd
                
                poweruse_prevd = Paesu_Record.objects.filter(user_id_p = request.user).get(date = yesterday).poweruse_today if Paesu_Record.objects.filter(user_id_p = request.user).filter(date = yesterday).exists() else 0
                poweruse_prevd = 0 if poweruse_prevd is None else poweruse_prevd

                context = {
                    "psrecord" : psrecord[0] if psrecord.count() != 0 else 0,
                    "date" : date,
                    "waterworks_prevd" : waterworks_prevd,
                    "underwater_prevd" : underwater_prevd,
                    "diswaste_prevd" : diswaste_prevd,
                    "poweruse_prevd" : poweruse_prevd,
                }

                return render(request, 'lists/insert.html', context)
            
            except:
                context = {
                    "date" : date,
                }
                return render(request, 'lists/insert.html', context)
        
    else:
        return redirect('/')


# For API views
# class DateExcel(APIView):

#     def post(self, request):
#         try:
#             filename = 'sample.xlsx'
#             wb = openpyxl.load_workbook(filename)

#             locale.setlocale(locale.LC_TIME, "ko_KR.UTF-8")

#             psrecord = Paesu_Record.objects.filter(user_id = User.objects.get(business_name = request.data['selected_corp'])).get(date = request.data['selected_date'])
#             date_obj = datetime.datetime.strptime(request.data['selected_date'], "%Y-%m-%d")
#             formatted_date_str = date_obj.strftime("%Y년%m월%d일")
            
#             # Cover title
#             wb['Cover']['B2'].value = request.data['selected_corp'] + ' ' +formatted_date_str + ' 레포트'

#             # Cover df
#             wb['Cover']['B5'].value = formatted_date_str
#             wb['Cover']['C5'].value = psrecord.diswaste_today
#             wb['Cover']['D5'].value = psrecord.diswaste_used
#             wb['Cover']['E5'].value = psrecord.poweruse_today
#             wb['Cover']['F5'].value = psrecord.poweruse_used

#             wb['0000-00-00']['D4'] = formatted_date_str

#             weekday_dict = {0: '월요일', 1: '화요일', 2: '수요일', 3: '목요일', 4: '금요일', 5: '토요일', 6: '일요일'}
#             day_of_week_kor_str = weekday_dict[date_obj.weekday()] 
#             wb['0000-00-00']['J4'] = day_of_week_kor_str
            
#             if psrecord.date_weather is not None: wb['0000-00-00']['P4'].value = psrecord.date_weather
#             if psrecord.date_temperature is not None: wb['0000-00-00']['W4'].value = psrecord.date_temperature

#             colorfill_dict = {0:'F', 1:'G', 2:'H', 3:'I', 4:'J', 5:'K', 6:'L', 7:'M', 8:'N', 9:'O', 10:'P', 11:'Q', 12:'R', 13:'S', 14:'T', 15:'U', 16:'V', 17:'W', 18:'X', 19:'Y', 20:'Z', 21:'AA', 22:'AB', 23:'AC', 24:'F'}
            
#             # 1. 운영 시간대 
#             for i in range(int((psrecord.op_start).split(":")[0]), int((psrecord.op_end).split(":")[0])+1):
#                 cn = colorfill_dict[i] + '8'
#                 wb['0000-00-00'][cn].fill = PatternFill(fill_type='solid',fgColor=Color('000000'))
            
#             # 1. 배출시설 가동(조업)시간대 
#             for i in range(int((psrecord.emission_start).split(":")[0]), int((psrecord.emission_end).split(":")[0])+1):
#                 cn = colorfill_dict[i] + '13'
#                 wb['0000-00-00'][cn].fill = PatternFill(fill_type='solid',fgColor=Color('000000'))

#             # 2. 방지시설 가동시간대(처리방법: 물리화학적처리)
#             for i in range(int((psrecord.prev_start).split(":")[0]), int((psrecord.prev_end).split(":")[0])+1):
#                 cn = colorfill_dict[i] + '18'
#                 wb['0000-00-00'][cn].fill = PatternFill(fill_type='solid',fgColor=Color('000000'))

#             # 3. 원료 또는 첨가제 등의 사용량
#             if psrecord.at_detergent is not None: wb['0000-00-00']['I23'].value = psrecord.at_detergent
#             if psrecord.at_wax is not None: wb['0000-00-00']['L23'].value = psrecord.at_wax
#             if psrecord.at_pom is not None: wb['0000-00-00']['O23'].value = psrecord.at_pom
#             if psrecord.at_sub1 is not None: wb['0000-00-00']['R22'].value = psrecord.at_sub1
#             if psrecord.at_sub1_memo is not None: wb['0000-00-00']['R23'].value = psrecord.at_sub1_memo
#             if psrecord.at_sub2 is not None: wb['0000-00-00']['U22'].value = psrecord.at_sub2
#             if psrecord.at_sub2_memo is not None: wb['0000-00-00']['U23'].value = psrecord.at_sub2_memo
#             if psrecord.at_sub3 is not None: wb['0000-00-00']['X22'].value = psrecord.at_sub3
#             if psrecord.at_sub3_memo is not None: wb['0000-00-00']['X23'].value = psrecord.at_sub3_memo
#             if psrecord.at_washnum is not None: wb['0000-00-00']['AA23'].value = psrecord.at_washnum

#             # 4. 용수공급원별 사용량과 폐수배출량
#             if psrecord.waterworks_prevd is not None: wb['0000-00-00']['E28'].value  = psrecord.waterworks_prevd
#             if psrecord.waterworks_used is not None: wb['0000-00-00']['H28'].value  = psrecord.waterworks_used
#             if psrecord.waterworks_today is not None: wb['0000-00-00']['K28'].value  = psrecord.waterworks_today
#             if psrecord.underwater_prevd is not None: wb['0000-00-00']['E29'].value  = psrecord.underwater_prevd
#             if psrecord.underwater_used is not None: wb['0000-00-00']['H29'].value  = psrecord.underwater_used
#             if psrecord.underwater_today is not None: wb['0000-00-00']['K29'].value  = psrecord.underwater_today

#             if psrecord.genwaster is not None: wb['0000-00-00']['W27'].value = psrecord.genwaster
#             # if psrecord.genwaster is not None: wb['0000-00-00']['W27'].value = psrecord.genwaster
#             if psrecord.diswaste_prevd is not None: wb['0000-00-00']['T28'].value = psrecord.diswaste_prevd
#             if psrecord.diswaste_used is not None: wb['0000-00-00']['W28'].value = psrecord.diswaste_used
#             if psrecord.diswaste_today is not None: wb['0000-00-00']['Z28'].value = psrecord.diswaste_today
#             if psrecord.reuse is not None: wb['0000-00-00']['W29'].value = psrecord.reuse
#             # if psrecord.reuse is not None: wb['0000-00-00']['W29'].value = psrecord.reuse 

#             # 5. 전력사용량
#             wb['0000-00-00']['B33'].value = int((psrecord.poweruse_end).split(":")[0]) - int((psrecord.poweruse_start).split(":")[0]) 
#             if psrecord.poweruse_used is not None: wb['0000-00-00']['F33'].value = psrecord.poweruse_used
#             if psrecord.poweruse_today is not None: wb['0000-00-00']['U33'].value = psrecord.poweruse_today
#             if psrecord.poweruse_etc is not None: wb['0000-00-00']['Z33'].value = psrecord.poweruse_etc
#             if (wb['0000-00-00']['W28'].value is None) and (wb['0000-00-00']['U33'].value is None): wb['0000-00-00']['J33'].value = int(wb['0000-00-00']['U33'].value) / int(wb['0000-00-00']['W28'].value)

#             # 6. 약품사용량
#             if psrecord.med1_name is not None: wb['0000-00-00']['B37'].value  = psrecord.med1_name
#             if psrecord.med1_buy is not None: wb['0000-00-00']['E37'].value  = psrecord.med1_buy
#             if psrecord.med1_used is not None: wb['0000-00-00']['H37'].value  = psrecord.med1_used
#             if psrecord.med1_balance is not None: wb['0000-00-00']['K37'].value  = psrecord.med1_balance
#             if psrecord.med1_etc is not None: wb['0000-00-00']['N37'].value  = psrecord.med1_etc
#             if psrecord.med2_name is not None: wb['0000-00-00']['P37'].value  = psrecord.med2_name
#             if psrecord.med2_buy is not None: wb['0000-00-00']['S37'].value  = psrecord.med2_buy
#             if psrecord.med2_used is not None: wb['0000-00-00']['V37'].value  = psrecord.med2_used
#             if psrecord.med2_balance is not None: wb['0000-00-00']['Y37'].value  = psrecord.med2_balance
#             if psrecord.med2_etc is not None: wb['0000-00-00']['AB37'].value  = psrecord.med2_etc
#             if psrecord.med3_name is not None: wb['0000-00-00']['B38'].value  = psrecord.med3_name
#             if psrecord.med3_buy is not None: wb['0000-00-00']['E38'].value  = psrecord.med3_buy
#             if psrecord.med3_used is not None: wb['0000-00-00']['H38'].value  = psrecord.med3_used
#             if psrecord.med3_balance is not None: wb['0000-00-00']['K38'].value  = psrecord.med3_balance
#             if psrecord.med3_etc is not None: wb['0000-00-00']['N38'].value  = psrecord.med3_etc
#             if psrecord.med4_name is not None: wb['0000-00-00']['P38'].value  = psrecord.med4_name
#             if psrecord.med4_buy is not None: wb['0000-00-00']['S38'].value  = psrecord.med4_buy
#             if psrecord.med4_used is not None: wb['0000-00-00']['V38'].value  = psrecord.med4_used
#             if psrecord.med4_balance is not None: wb['0000-00-00']['Y38'].value  = psrecord.med4_balance
#             if psrecord.med4_etc is not None: wb['0000-00-00']['AB38'].value  = psrecord.med4_etc
        
#             # 7. 슬러지처리시설
#             if psrecord.sluge_gene is not None: wb['0000-00-00']['B42'].value  = psrecord.waterworks_prevd
#             if psrecord.sluge_used is not None: wb['0000-00-00']['I42'].value  = psrecord.waterworks_prevd
#             if psrecord.sluge_keep is not None: wb['0000-00-00']['N42'].value  = psrecord.sluge_keep
#             if psrecord.sluge_func is not None: wb['0000-00-00']['S42'].value  = psrecord.sluge_func
#             if psrecord.sluge_place is not None: wb['0000-00-00']['Y42'].value  = psrecord.sluge_place
#             if psrecord.sluge_selfplace is not None: wb['0000-00-00']['S42'].value  = psrecord.sluge_selfplace
#             if psrecord.sluge_corp is not None: wb['0000-00-00']['Y42'].value  = psrecord.sluge_corp

#             # 8. 방지시설 고장유무 및 특기사항
#             if psrecord.remarks is not None: wb['0000-00-00']['H43'].value  = psrecord.remarks

#             # 9. 지도 · 점검 받은 사항
#             if psrecord.advise is not None: wb['0000-00-00']['T43'].value  = psrecord.advise

#             # Change Index1 sheet name.
#             wb['0000-00-00'].title = request.data['selected_date']

#             file_obj = io.BytesIO()
#             wb.save(file_obj)
#             file_obj.seek(0)

#             # response = FileResponse(file_obj, as_attachment=True, filename='example.xlsx')
#             # response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#             response = HttpResponse(file_obj,content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#             response['Content-Disposition'] = 'attachment; filename=tttt.xlsx'
#             return response
        
#         except:
#             return Response({'message': 'Bad Request'}, status=status.HTTP_400_BAD_REQUEST)


def DownloadExcel(request):
    try:
        filename = 'sample.xlsx'
        wb = openpyxl.load_workbook(filename)
        locale.setlocale(locale.LC_TIME, "ko_KR.UTF-8")

        psrecord = Paesu_Record.objects.filter(user_id_p = User.objects.get(business_name = request.GET['selected_corp'])).get(date = request.GET['selected_date'])
        date_obj = datetime.datetime.strptime(request.GET['selected_date'], "%Y-%m-%d")
        formatted_date_str = date_obj.strftime("%Y년%m월%d일")
        
        # Cover title
        wb['Cover']['B2'].value = request.GET['selected_corp'] + ' ' +formatted_date_str + ' 레포트'

        # Cover df
        wb['Cover']['B5'].value = formatted_date_str
        wb['Cover']['C5'].value = psrecord.diswaste_today
        wb['Cover']['D5'].value = psrecord.diswaste_used
        wb['Cover']['E5'].value = psrecord.poweruse_today
        wb['Cover']['F5'].value = psrecord.poweruse_used
        wb['0000-00-00']['D4'] = formatted_date_str
        weekday_dict = {0: '월요일', 1: '화요일', 2: '수요일', 3: '목요일', 4: '금요일', 5: '토요일', 6: '일요일'}
        day_of_week_kor_str = weekday_dict[date_obj.weekday()] 
        wb['0000-00-00']['J4'] = day_of_week_kor_str
        
        if psrecord.date_weather is not None: wb['0000-00-00']['P4'].value = psrecord.date_weather
        if psrecord.date_temperature is not None: wb['0000-00-00']['W4'].value = psrecord.date_temperature
        colorfill_dict = {0:'F', 1:'G', 2:'H', 3:'I', 4:'J', 5:'K', 6:'L', 7:'M', 8:'N', 9:'O', 10:'P', 11:'Q', 12:'R', 13:'S', 14:'T', 15:'U', 16:'V', 17:'W', 18:'X', 19:'Y', 20:'Z', 21:'AA', 22:'AB', 23:'AC', 24:'F'}
        
        # 1. 운영 시간대 
        for i in range(int((psrecord.op_start).split(":")[0]), int((psrecord.op_end).split(":")[0])+1):
            cn = colorfill_dict[i] + '8'
            wb['0000-00-00'][cn].fill = PatternFill(fill_type='solid',fgColor=Color('000000'))
        
        # 1. 배출시설 가동(조업)시간대 
        for i in range(int((psrecord.emission_start).split(":")[0]), int((psrecord.emission_end).split(":")[0])+1):
            cn = colorfill_dict[i] + '13'
            wb['0000-00-00'][cn].fill = PatternFill(fill_type='solid',fgColor=Color('000000'))

        # 2. 방지시설 가동시간대(처리방법: 물리화학적처리)
        for i in range(int((psrecord.prev_start).split(":")[0]), int((psrecord.prev_end).split(":")[0])+1):
            cn = colorfill_dict[i] + '18'
            wb['0000-00-00'][cn].fill = PatternFill(fill_type='solid',fgColor=Color('000000'))

        # 3. 원료 또는 첨가제 등의 사용량
        if psrecord.at_detergent is not None: wb['0000-00-00']['I23'].value = psrecord.at_detergent
        if psrecord.at_wax is not None: wb['0000-00-00']['L23'].value = psrecord.at_wax
        if psrecord.at_pom is not None: wb['0000-00-00']['O23'].value = psrecord.at_pom
        if psrecord.at_sub1 is not None: wb['0000-00-00']['R22'].value = psrecord.at_sub1
        if psrecord.at_sub1_memo is not None: wb['0000-00-00']['R23'].value = psrecord.at_sub1_memo
        if psrecord.at_sub2 is not None: wb['0000-00-00']['U22'].value = psrecord.at_sub2
        if psrecord.at_sub2_memo is not None: wb['0000-00-00']['U23'].value = psrecord.at_sub2_memo
        if psrecord.at_sub3 is not None: wb['0000-00-00']['X22'].value = psrecord.at_sub3
        if psrecord.at_sub3_memo is not None: wb['0000-00-00']['X23'].value = psrecord.at_sub3_memo
        if psrecord.at_washnum is not None: wb['0000-00-00']['AA23'].value = psrecord.at_washnum

        # 4. 용수공급원별 사용량과 폐수배출량
        if psrecord.waterworks_prevd is not None: wb['0000-00-00']['E28'].value  = psrecord.waterworks_prevd
        if psrecord.waterworks_used is not None: wb['0000-00-00']['H28'].value  = psrecord.waterworks_used
        if psrecord.waterworks_today is not None: wb['0000-00-00']['K28'].value  = psrecord.waterworks_today
        if psrecord.underwater_prevd is not None: wb['0000-00-00']['E29'].value  = psrecord.underwater_prevd
        if psrecord.underwater_used is not None: wb['0000-00-00']['H29'].value  = psrecord.underwater_used
        if psrecord.underwater_today is not None: wb['0000-00-00']['K29'].value  = psrecord.underwater_today
        if psrecord.genwaster is not None: wb['0000-00-00']['W27'].value = psrecord.genwaster

        # if psrecord.genwaster is not None: wb['0000-00-00']['W27'].value = psrecord.genwaster
        if psrecord.diswaste_prevd is not None: wb['0000-00-00']['T28'].value = psrecord.diswaste_prevd
        if psrecord.diswaste_used is not None: wb['0000-00-00']['W28'].value = psrecord.diswaste_used
        if psrecord.diswaste_today is not None: wb['0000-00-00']['Z28'].value = psrecord.diswaste_today
        if psrecord.reuse is not None: wb['0000-00-00']['W29'].value = psrecord.reuse
        # if psrecord.reuse is not None: wb['0000-00-00']['W29'].value = psrecord.reuse
        #  
        # 5. 전력사용량
        wb['0000-00-00']['B33'].value = int((psrecord.poweruse_end).split(":")[0]) - int((psrecord.poweruse_start).split(":")[0]) 
        if psrecord.poweruse_used is not None: wb['0000-00-00']['F33'].value = psrecord.poweruse_used
        if psrecord.poweruse_today is not None: wb['0000-00-00']['U33'].value = psrecord.poweruse_today
        if psrecord.poweruse_etc is not None: wb['0000-00-00']['Z33'].value = psrecord.poweruse_etc
        if (wb['0000-00-00']['W28'].value is None) and (wb['0000-00-00']['U33'].value is None): wb['0000-00-00']['J33'].value = int(wb['0000-00-00']['U33'].value) / int(wb['0000-00-00']['W28'].value)

        # 6. 약품사용량
        if psrecord.med1_name is not None: wb['0000-00-00']['B37'].value  = psrecord.med1_name
        if psrecord.med1_buy is not None: wb['0000-00-00']['E37'].value  = psrecord.med1_buy
        if psrecord.med1_used is not None: wb['0000-00-00']['H37'].value  = psrecord.med1_used
        if psrecord.med1_balance is not None: wb['0000-00-00']['K37'].value  = psrecord.med1_balance
        if psrecord.med1_etc is not None: wb['0000-00-00']['N37'].value  = psrecord.med1_etc
        if psrecord.med2_name is not None: wb['0000-00-00']['P37'].value  = psrecord.med2_name
        if psrecord.med2_buy is not None: wb['0000-00-00']['S37'].value  = psrecord.med2_buy
        if psrecord.med2_used is not None: wb['0000-00-00']['V37'].value  = psrecord.med2_used
        if psrecord.med2_balance is not None: wb['0000-00-00']['Y37'].value  = psrecord.med2_balance
        if psrecord.med2_etc is not None: wb['0000-00-00']['AB37'].value  = psrecord.med2_etc
        if psrecord.med3_name is not None: wb['0000-00-00']['B38'].value  = psrecord.med3_name
        if psrecord.med3_buy is not None: wb['0000-00-00']['E38'].value  = psrecord.med3_buy
        if psrecord.med3_used is not None: wb['0000-00-00']['H38'].value  = psrecord.med3_used
        if psrecord.med3_balance is not None: wb['0000-00-00']['K38'].value  = psrecord.med3_balance
        if psrecord.med3_etc is not None: wb['0000-00-00']['N38'].value  = psrecord.med3_etc
        if psrecord.med4_name is not None: wb['0000-00-00']['P38'].value  = psrecord.med4_name
        if psrecord.med4_buy is not None: wb['0000-00-00']['S38'].value  = psrecord.med4_buy
        if psrecord.med4_used is not None: wb['0000-00-00']['V38'].value  = psrecord.med4_used
        if psrecord.med4_balance is not None: wb['0000-00-00']['Y38'].value  = psrecord.med4_balance
        if psrecord.med4_etc is not None: wb['0000-00-00']['AB38'].value  = psrecord.med4_etc
    
        # 7. 슬러지처리시설
        if psrecord.sluge_gene is not None: wb['0000-00-00']['B42'].value  = psrecord.waterworks_prevd
        if psrecord.sluge_used is not None: wb['0000-00-00']['I42'].value  = psrecord.waterworks_prevd
        if psrecord.sluge_keep is not None: wb['0000-00-00']['N42'].value  = psrecord.sluge_keep
        if psrecord.sluge_func is not None: wb['0000-00-00']['S42'].value  = psrecord.sluge_func
        if psrecord.sluge_place is not None: wb['0000-00-00']['Y42'].value  = psrecord.sluge_place
        if psrecord.sluge_selfplace is not None: wb['0000-00-00']['S42'].value  = psrecord.sluge_selfplace
        if psrecord.sluge_corp is not None: wb['0000-00-00']['Y42'].value  = psrecord.sluge_corp

        # 8. 방지시설 고장유무 및 특기사항
        if psrecord.remarks is not None: wb['0000-00-00']['H43'].value  = psrecord.remarks

        # 9. 지도 · 점검 받은 사항
        if psrecord.advise is not None: wb['0000-00-00']['T43'].value  = psrecord.advise

        # Change Index1 sheet name.
        wb['0000-00-00'].title = request.GET['selected_date']
        file_obj = io.BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        response = FileResponse(file_obj, as_attachment=True, filename='example.xlsx')
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        return response
    
    except: return HttpResponse(status=204)


def DownloadExcelAll(request):
    print('here')

    print(request.GET)



    try:
        filename = 'sample.xlsx'
        wb = openpyxl.load_workbook(filename)
        locale.setlocale(locale.LC_TIME, "ko_KR.UTF-8")

        psrecord = Paesu_Record.objects.filter(user_id_p = User.objects.get(business_name = request.GET['selected_corp'])).filter(date__gte = request.GET['start_date'], date__lte = request.GET['end_date'])
        
        start_date_obj = datetime.datetime.strptime(request.GET['start_date'], "%Y-%m-%d")
        start_formatted_date_str = start_date_obj.strftime("%Y년%m월%d일")
        end_date_obj = datetime.datetime.strptime(request.GET['end_date'], "%Y-%m-%d")
        end_formatted_date_str = end_date_obj.strftime("%Y년%m월%d일")

        # Cover title
        wb['Cover']['B2'].value = request.GET['selected_corp'] + ' ' + start_formatted_date_str + ' ~ ' + end_formatted_date_str +' 레포트'
        start_row = 5

        for psn in range(len(psrecord)):
            if psn == 0:
                wb['Cover']['B5'].value = (psrecord[0].date).strftime("%Y년%m월%d일")
                wb['Cover']['C5'].value = psrecord[0].diswaste_today
                wb['Cover']['D5'].value = psrecord[0].diswaste_used
                wb['Cover']['E5'].value = psrecord[0].poweruse_today
                wb['Cover']['F5'].value = psrecord[0].poweruse_used

                weekday_dict = {0: '월요일', 1: '화요일', 2: '수요일', 3: '목요일', 4: '금요일', 5: '토요일', 6: '일요일'}
                day_of_week_kor_str = weekday_dict[psrecord[0].date.weekday()]
                wb['0000-00-00']['J4'] = day_of_week_kor_str

                if psrecord[0].date_weather is not None: wb['0000-00-00']['P4'].value = psrecord[0].date_weather
                if psrecord[0].date_temperature is not None: wb['0000-00-00']['W4'].value = psrecord[0].date_temperature

                colorfill_dict = {0:'F', 1:'G', 2:'H', 3:'I', 4:'J', 5:'K', 6:'L', 7:'M', 8:'N', 9:'O', 10:'P', 11:'Q', 12:'R', 13:'S', 14:'T', 15:'U', 16:'V', 17:'W', 18:'X', 19:'Y', 20:'Z', 21:'AA', 22:'AB', 23:'AC', 24:'F'}

                # 1. 운영 시간대 
                for i in range(int((psrecord[0].op_start).split(":")[0]), int((psrecord[0].op_end).split(":")[0])+1):
                    cn = colorfill_dict[i] + '8'
                    wb['0000-00-00'][cn].fill = PatternFill(fill_type='solid',fgColor=Color('000000'))

                # 1. 배출시설 가동(조업)시간대 
                for i in range(int((psrecord[0].emission_start).split(":")[0]), int((psrecord[0].emission_end).split(":")[0])+1):
                    cn = colorfill_dict[i] + '13'
                    wb['0000-00-00'][cn].fill = PatternFill(fill_type='solid',fgColor=Color('000000'))

                # 2. 방지시설 가동시간대(처리방법: 물리화학적처리)
                for i in range(int((psrecord[0].prev_start).split(":")[0]), int((psrecord[0].prev_end).split(":")[0])+1):
                    cn = colorfill_dict[i] + '18'
                    wb['0000-00-00'][cn].fill = PatternFill(fill_type='solid',fgColor=Color('000000'))

                # 3. 원료 또는 첨가제 등의 사용량
                if psrecord[0].at_detergent is not None: wb['0000-00-00']['I23'].value = psrecord[0].at_detergent
                if psrecord[0].at_wax is not None: wb['0000-00-00']['L23'].value = psrecord[0].at_wax
                if psrecord[0].at_pom is not None: wb['0000-00-00']['O23'].value = psrecord[0].at_pom
                if psrecord[0].at_sub1 is not None: wb['0000-00-00']['R22'].value = psrecord[0].at_sub1
                if psrecord[0].at_sub1_memo is not None: wb['0000-00-00']['R23'].value = psrecord[0].at_sub1_memo
                if psrecord[0].at_sub2 is not None: wb['0000-00-00']['U22'].value = psrecord[0].at_sub2
                if psrecord[0].at_sub2_memo is not None: wb['0000-00-00']['U23'].value = psrecord[0].at_sub2_memo
                if psrecord[0].at_sub3 is not None: wb['0000-00-00']['X22'].value = psrecord[0].at_sub3
                if psrecord[0].at_sub3_memo is not None: wb['0000-00-00']['X23'].value = psrecord[0].at_sub3_memo
                if psrecord[0].at_washnum is not None: wb['0000-00-00']['AA23'].value = psrecord[0].at_washnum

                # 4. 용수공급원별 사용량과 폐수배출량
                if psrecord[0].waterworks_prevd is not None: wb['0000-00-00']['E28'].value  = psrecord[0].waterworks_prevd
                if psrecord[0].waterworks_used is not None: wb['0000-00-00']['H28'].value  = psrecord[0].waterworks_used
                if psrecord[0].waterworks_today is not None: wb['0000-00-00']['K28'].value  = psrecord[0].waterworks_today
                if psrecord[0].underwater_prevd is not None: wb['0000-00-00']['E29'].value  = psrecord[0].underwater_prevd
                if psrecord[0].underwater_used is not None: wb['0000-00-00']['H29'].value  = psrecord[0].underwater_used
                if psrecord[0].underwater_today is not None: wb['0000-00-00']['K29'].value  = psrecord[0].underwater_today
                if psrecord[0].genwaster is not None: wb['0000-00-00']['W27'].value = psrecord[0].genwaster

                # if psrecord.genwaster is not None: wb['0000-00-00']['W27'].value = psrecord.genwaster
                if psrecord[0].diswaste_prevd is not None: wb['0000-00-00']['T28'].value = psrecord[0].diswaste_prevd
                if psrecord[0].diswaste_used is not None: wb['0000-00-00']['W28'].value = psrecord[0].diswaste_used
                if psrecord[0].diswaste_today is not None: wb['0000-00-00']['Z28'].value = psrecord[0].diswaste_today
                if psrecord[0].reuse is not None: wb['0000-00-00']['W29'].value = psrecord[0].reuse
                # if psrecord.reuse is not None: wb['0000-00-00']['W29'].value = psrecord.reuse
                #  
                # 5. 전력사용량
                wb['0000-00-00']['B33'].value = int((psrecord[0].poweruse_end).split(":")[0]) - int((psrecord[0].poweruse_start).split(":")[0]) 
                if psrecord[0].poweruse_used is not None: wb['0000-00-00']['F33'].value = psrecord[0].poweruse_used
                if psrecord[0].poweruse_today is not None: wb['0000-00-00']['U33'].value = psrecord[0].poweruse_today
                if psrecord[0].poweruse_etc is not None: wb['0000-00-00']['Z33'].value = psrecord[0].poweruse_etc
                if (wb['0000-00-00']['W28'].value is None) and (wb['0000-00-00']['U33'].value is None): wb['0000-00-00']['J33'].value = int(wb['0000-00-00']['U33'].value) / int(wb['0000-00-00']['W28'].value)

                # 6. 약품사용량
                if psrecord[0].med1_name is not None: wb['0000-00-00']['B37'].value  = psrecord[0].med1_name
                if psrecord[0].med1_buy is not None: wb['0000-00-00']['E37'].value  = psrecord[0].med1_buy
                if psrecord[0].med1_used is not None: wb['0000-00-00']['H37'].value  = psrecord[0].med1_used
                if psrecord[0].med1_balance is not None: wb['0000-00-00']['K37'].value  = psrecord[0].med1_balance
                if psrecord[0].med1_etc is not None: wb['0000-00-00']['N37'].value  = psrecord[0].med1_etc
                if psrecord[0].med2_name is not None: wb['0000-00-00']['P37'].value  = psrecord[0].med2_name
                if psrecord[0].med2_buy is not None: wb['0000-00-00']['S37'].value  = psrecord[0].med2_buy
                if psrecord[0].med2_used is not None: wb['0000-00-00']['V37'].value  = psrecord[0].med2_used
                if psrecord[0].med2_balance is not None: wb['0000-00-00']['Y37'].value  = psrecord[0].med2_balance
                if psrecord[0].med2_etc is not None: wb['0000-00-00']['AB37'].value  = psrecord[0].med2_etc
                if psrecord[0].med3_name is not None: wb['0000-00-00']['B38'].value  = psrecord[0].med3_name
                if psrecord[0].med3_buy is not None: wb['0000-00-00']['E38'].value  = psrecord[0].med3_buy
                if psrecord[0].med3_used is not None: wb['0000-00-00']['H38'].value  = psrecord[0].med3_used
                if psrecord[0].med3_balance is not None: wb['0000-00-00']['K38'].value  = psrecord[0].med3_balance
                if psrecord[0].med3_etc is not None: wb['0000-00-00']['N38'].value  = psrecord[0].med3_etc
                if psrecord[0].med4_name is not None: wb['0000-00-00']['P38'].value  = psrecord[0].med4_name
                if psrecord[0].med4_buy is not None: wb['0000-00-00']['S38'].value  = psrecord[0].med4_buy
                if psrecord[0].med4_used is not None: wb['0000-00-00']['V38'].value  = psrecord[0].med4_used
                if psrecord[0].med4_balance is not None: wb['0000-00-00']['Y38'].value  = psrecord[0].med4_balance
                if psrecord[0].med4_etc is not None: wb['0000-00-00']['AB38'].value  = psrecord[0].med4_etc

                # 7. 슬러지처리시설
                if psrecord[0].sluge_gene is not None: wb['0000-00-00']['B42'].value  = psrecord[0].waterworks_prevd
                if psrecord[0].sluge_used is not None: wb['0000-00-00']['I42'].value  = psrecord[0].waterworks_prevd
                if psrecord[0].sluge_keep is not None: wb['0000-00-00']['N42'].value  = psrecord[0].sluge_keep
                if psrecord[0].sluge_func is not None: wb['0000-00-00']['S42'].value  = psrecord[0].sluge_func
                if psrecord[0].sluge_place is not None: wb['0000-00-00']['Y42'].value  = psrecord[0].sluge_place
                if psrecord[0].sluge_selfplace is not None: wb['0000-00-00']['S42'].value  = psrecord[0].sluge_selfplace
                if psrecord[0].sluge_corp is not None: wb['0000-00-00']['Y42'].value  = psrecord[0].sluge_corp

                # 8. 방지시설 고장유무 및 특기사항
                if psrecord[0].remarks is not None: wb['0000-00-00']['H43'].value  = psrecord[0].remarks

                # 9. 지도 · 점검 받은 사항
                if psrecord[0].advise is not None: wb['0000-00-00']['T43'].value  = psrecord[0].advise

                wb['0000-00-00'].title = str(psrecord[0].date)
                copy_Sheet = wb.copy_worksheet(wb[str(psrecord[0].date)])
                copy_Sheet.title = '0000-00-00'
                wb['0000-00-00'].sheet_view.showGridLines = False


            else:
                wb['Cover']['B'+str(start_row+psn)].value = (psrecord[psn].date).strftime("%Y년%m월%d일")
                wb['Cover']['C'+str(start_row+psn)].value = psrecord[psn].diswaste_today
                wb['Cover']['D'+str(start_row+psn)].value = psrecord[psn].diswaste_used
                wb['Cover']['E'+str(start_row+psn)].value = psrecord[psn].poweruse_today
                wb['Cover']['F'+str(start_row+psn)].value = psrecord[psn].poweruse_used

                colorfill_dict = {0:'F', 1:'G', 2:'H', 3:'I', 4:'J', 5:'K', 6:'L', 7:'M', 8:'N', 9:'O', 10:'P', 11:'Q', 12:'R', 13:'S', 14:'T', 15:'U', 16:'V', 17:'W', 18:'X', 19:'Y', 20:'Z', 21:'AA', 22:'AB', 23:'AC', 24:'F'}

                # 1. 운영 시간대 
                for i in range(int((psrecord[psn].op_start).split(":")[0]), int((psrecord[psn].op_end).split(":")[0])+1):
                    cn = colorfill_dict[i] + '8'
                    wb['0000-00-00'][cn].fill = PatternFill(fill_type='solid',fgColor=Color('000000'))

                # 1. 배출시설 가동(조업)시간대 
                for i in range(int((psrecord[psn].emission_start).split(":")[0]), int((psrecord[psn].emission_end).split(":")[0])+1):
                    cn = colorfill_dict[i] + '13'
                    wb['0000-00-00'][cn].fill = PatternFill(fill_type='solid',fgColor=Color('000000'))

                # 2. 방지시설 가동시간대(처리방법: 물리화학적처리)
                for i in range(int((psrecord[psn].prev_start).split(":")[0]), int((psrecord[psn].prev_end).split(":")[0])+1):
                    cn = colorfill_dict[i] + '18'
                    wb['0000-00-00'][cn].fill = PatternFill(fill_type='solid',fgColor=Color('000000'))

                # 3. 원료 또는 첨가제 등의 사용량
                if psrecord[psn].at_detergent is not None: wb['0000-00-00']['I23'].value = psrecord[psn].at_detergent
                if psrecord[psn].at_wax is not None: wb['0000-00-00']['L23'].value = psrecord[psn].at_wax
                if psrecord[psn].at_pom is not None: wb['0000-00-00']['O23'].value = psrecord[psn].at_pom
                if psrecord[psn].at_sub1 is not None: wb['0000-00-00']['R22'].value = psrecord[psn].at_sub1
                if psrecord[psn].at_sub1_memo is not None: wb['0000-00-00']['R23'].value = psrecord[psn].at_sub1_memo
                if psrecord[psn].at_sub2 is not None: wb['0000-00-00']['U22'].value = psrecord[psn].at_sub2
                if psrecord[psn].at_sub2_memo is not None: wb['0000-00-00']['U23'].value = psrecord[psn].at_sub2_memo
                if psrecord[psn].at_sub3 is not None: wb['0000-00-00']['X22'].value = psrecord[psn].at_sub3
                if psrecord[psn].at_sub3_memo is not None: wb['0000-00-00']['X23'].value = psrecord[psn].at_sub3_memo
                if psrecord[psn].at_washnum is not None: wb['0000-00-00']['AA23'].value = psrecord[psn].at_washnum

                # 4. 용수공급원별 사용량과 폐수배출량
                if psrecord[psn].waterworks_prevd is not None: wb['0000-00-00']['E28'].value  = psrecord[psn].waterworks_prevd
                if psrecord[psn].waterworks_used is not None: wb['0000-00-00']['H28'].value  = psrecord[psn].waterworks_used
                if psrecord[psn].waterworks_today is not None: wb['0000-00-00']['K28'].value  = psrecord[psn].waterworks_today
                if psrecord[psn].underwater_prevd is not None: wb['0000-00-00']['E29'].value  = psrecord[psn].underwater_prevd
                if psrecord[psn].underwater_used is not None: wb['0000-00-00']['H29'].value  = psrecord[psn].underwater_used
                if psrecord[psn].underwater_today is not None: wb['0000-00-00']['K29'].value  = psrecord[psn].underwater_today
                if psrecord[psn].genwaster is not None: wb['0000-00-00']['W27'].value = psrecord[psn].genwaster

                # if psrecord.genwaster is not None: wb['0000-00-00']['W27'].value = psrecord.genwaster
                if psrecord[psn].diswaste_prevd is not None: wb['0000-00-00']['T28'].value = psrecord[psn].diswaste_prevd
                if psrecord[psn].diswaste_used is not None: wb['0000-00-00']['W28'].value = psrecord[psn].diswaste_used
                if psrecord[psn].diswaste_today is not None: wb['0000-00-00']['Z28'].value = psrecord[psn].diswaste_today
                if psrecord[psn].reuse is not None: wb['0000-00-00']['W29'].value = psrecord[psn].reuse
                # if psrecord.reuse is not None: wb['0000-00-00']['W29'].value = psrecord.reuse
                # #  
                # 5. 전력사용량
                wb['0000-00-00']['B33'].value = int((psrecord[psn].poweruse_end).split(":")[0]) - int((psrecord[psn].poweruse_start).split(":")[0]) 
                if psrecord[psn].poweruse_used is not None: wb['0000-00-00']['F33'].value = psrecord[psn].poweruse_used
                if psrecord[psn].poweruse_today is not None: wb['0000-00-00']['U33'].value = psrecord[psn].poweruse_today
                if psrecord[psn].poweruse_etc is not None: wb['0000-00-00']['Z33'].value = psrecord[psn].poweruse_etc
                if (wb['0000-00-00']['W28'].value is None) and (wb['0000-00-00']['U33'].value is None): wb['0000-00-00']['J33'].value = int(wb['0000-00-00']['U33'].value) / int(wb['0000-00-00']['W28'].value)

                # 6. 약품사용량
                if psrecord[psn].med1_name is not None: wb['0000-00-00']['B37'].value  = psrecord[psn].med1_name
                if psrecord[psn].med1_buy is not None: wb['0000-00-00']['E37'].value  = psrecord[psn].med1_buy
                if psrecord[psn].med1_used is not None: wb['0000-00-00']['H37'].value  = psrecord[psn].med1_used
                if psrecord[psn].med1_balance is not None: wb['0000-00-00']['K37'].value  = psrecord[psn].med1_balance
                if psrecord[psn].med1_etc is not None: wb['0000-00-00']['N37'].value  = psrecord[psn].med1_etc
                if psrecord[psn].med2_name is not None: wb['0000-00-00']['P37'].value  = psrecord[psn].med2_name
                if psrecord[psn].med2_buy is not None: wb['0000-00-00']['S37'].value  = psrecord[psn].med2_buy
                if psrecord[psn].med2_used is not None: wb['0000-00-00']['V37'].value  = psrecord[psn].med2_used
                if psrecord[psn].med2_balance is not None: wb['0000-00-00']['Y37'].value  = psrecord[psn].med2_balance
                if psrecord[psn].med2_etc is not None: wb['0000-00-00']['AB37'].value  = psrecord[psn].med2_etc
                if psrecord[psn].med3_name is not None: wb['0000-00-00']['B38'].value  = psrecord[psn].med3_name
                if psrecord[psn].med3_buy is not None: wb['0000-00-00']['E38'].value  = psrecord[psn].med3_buy
                if psrecord[psn].med3_used is not None: wb['0000-00-00']['H38'].value  = psrecord[psn].med3_used
                if psrecord[psn].med3_balance is not None: wb['0000-00-00']['K38'].value  = psrecord[psn].med3_balance
                if psrecord[psn].med3_etc is not None: wb['0000-00-00']['N38'].value  = psrecord[psn].med3_etc
                if psrecord[psn].med4_name is not None: wb['0000-00-00']['P38'].value  = psrecord[psn].med4_name
                if psrecord[psn].med4_buy is not None: wb['0000-00-00']['S38'].value  = psrecord[psn].med4_buy
                if psrecord[psn].med4_used is not None: wb['0000-00-00']['V38'].value  = psrecord[psn].med4_used
                if psrecord[psn].med4_balance is not None: wb['0000-00-00']['Y38'].value  = psrecord[psn].med4_balance
                if psrecord[psn].med4_etc is not None: wb['0000-00-00']['AB38'].value  = psrecord[psn].med4_etc

                # 7. 슬러지처리시설
                if psrecord[psn].sluge_gene is not None: wb['0000-00-00']['B42'].value  = psrecord[psn].waterworks_prevd
                if psrecord[psn].sluge_used is not None: wb['0000-00-00']['I42'].value  = psrecord[psn].waterworks_prevd
                if psrecord[psn].sluge_keep is not None: wb['0000-00-00']['N42'].value  = psrecord[psn].sluge_keep
                if psrecord[psn].sluge_func is not None: wb['0000-00-00']['S42'].value  = psrecord[psn].sluge_func
                if psrecord[psn].sluge_place is not None: wb['0000-00-00']['Y42'].value  = psrecord[psn].sluge_place
                if psrecord[psn].sluge_selfplace is not None: wb['0000-00-00']['S42'].value  = psrecord[psn].sluge_selfplace
                if psrecord[psn].sluge_corp is not None: wb['0000-00-00']['Y42'].value  = psrecord[psn].sluge_corp

                # 8. 방지시설 고장유무 및 특기사항
                if psrecord[psn].remarks is not None: wb['0000-00-00']['H43'].value  = psrecord[psn].remarks

                # 9. 지도 · 점검 받은 사항
                if psrecord[psn].advise is not None: wb['0000-00-00']['T43'].value  = psrecord[psn].advise

                wb['0000-00-00'].title = str(psrecord[psn].date)
                copy_Sheet = wb.copy_worksheet(wb[str(psrecord[psn].date)])
                copy_Sheet.title = '0000-00-00'
                wb['0000-00-00'].sheet_view.showGridLines = False


            
        del wb['0000-00-00']
        file_obj = io.BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        response = FileResponse(file_obj, as_attachment=True, filename='example.xlsx')
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        return response
    
    except: return HttpResponse(status=204)